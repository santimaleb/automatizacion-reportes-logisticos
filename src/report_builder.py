# src/report_builder.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.config import DATA_OUTPUT_DIR

def build_excel_report(df, kpis, recommendations):
    """
    Construye el reporte en Excel con formato profesional
    """
    print("📋 Construyendo reporte en Excel...")
    
    filepath = f"{DATA_OUTPUT_DIR}/reporte_mensual.xlsx"
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Hoja 1: Resumen Ejecutivo
        summary_data = {
            'Métrica': [
                'Costo Total',
                'Costo Combustible',
                'Costo Peajes',
                'Costo Personal',
                'Total Km',
                'Total Horas',
                'Costo por Km',
                'Costo por Hora'
            ],
            'Valor': [
                f"${kpis['costo_total']:,.2f}",
                f"${kpis['costo_total_combustible']:,.2f}",
                f"${kpis['costo_total_peajes']:,.2f}",
                f"${kpis['costo_total_personal']:,.2f}",
                f"{kpis['total_km']:,.2f} km",
                f"{kpis['total_horas']:,.2f} hrs",
                f"${kpis['costo_por_km']:.2f}",
                f"${kpis['costo_por_hora']:.2f}"
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Hoja 2: KPIs por Ruta
        if 'kpis_por_ruta' in kpis:
            kpis['kpis_por_ruta'].to_excel(writer, sheet_name='KPIs por Ruta', index=False)
        
        # Hoja 3: Top Clientes
        if 'top_clientes' in kpis:
            kpis['top_clientes'].to_excel(writer, sheet_name='Top Clientes')
        
        # Hoja 4: Tendencia Mensual
        if 'tendencia_mensual' in kpis:
            kpis['tendencia_mensual'].to_excel(writer, sheet_name='Tendencia Mensual', index=False)
        
        # Hoja 5: Datos Completos
        df.to_excel(writer, sheet_name='Datos Completos', index=False)
        
        # Hoja 6: Recomendaciones
        if recommendations:
            rec_df = pd.DataFrame({'Recomendación': recommendations})
            rec_df.to_excel(writer, sheet_name='Recomendaciones', index=False)
    
    apply_excel_formatting(filepath)
    
    print(f"✅ Reporte guardado: {filepath}")
    return filepath

def apply_excel_formatting(filepath):
    """
    Aplica formato profesional al archivo Excel
    """
    wb = load_workbook(filepath)
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    print("   ✅ Formato aplicado al Excel")